#!/usr/bin/python
# -*- coding: UTF-8 -*-

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string
from operator import eq
from openpyxl.utils import get_column_letter


def get_max_rows(f1, f2, i1):
    """
    获取最大行数
    :param f1:主文件路径
    :param f2: 副文件路径
    :param i1: 主文件的sheet表index
    :param i2: 副文件的sheet表index
    :return: max_row：最大行数
             wb1：读取的主文件
             sheet1：读取的sheet表
             list(sheet1.merage_cells):文件的全部合并单元格列表
    """
    wb1 = openpyxl.load_workbook(f1)
    wb2 = openpyxl.load_workbook(f2, read_only=True)
    sheet1 = wb1.worksheets[i1]
    return wb1, sheet1, list(sheet1.merged_cells)


def check_PermissionError(path):
    """
    通过文件是否可以打开 判断文件是否打开
    :param path: 文件路径
    :return: None
    """
    try:
        f = open(path, "a")
    except OSError as e:
        return False
    else:
        f.close()
        return True


def fill_cols(lst1, lst2):
    """
    处理读取的列数据，长度不一问题
    :param lst1: 主文件读取的一行数据
    :param lst2: 副文件读取的一行数据
    :return: 处理完的lst1,lst2,最大列表长度
    """
    len1 = len(lst1)
    len2 = len(lst2)
    if len1 > len2:
        for i in range(len1 - len2):
            lst2.append("~\\")  # 填充特殊字符
        return lst1, lst2, len1
    elif len1 < len2:
        for i in range(len2 - len1):
            lst1.append("~\\")
        return lst1, lst2, len2
    else:
        return lst1, lst2, len1


def fill_rows(df1, df2,df1_rows,df2_rows):
    """
    处理读取的 dataframe行数不一的问题
    :param df1: 主文件读取的dataframe
    :param df2: 副文件读取的dataframe
    :param df1_rows: 主文件的最大行数
    :param df2_rows: 副文件的最大行数
    :return: 处理完的df1,df2
    """
    df1_cols,df2_cols=df1.shape[1],df2.shape[1]
    if df1_rows > df2_rows:
        n = df1_rows - df2_rows # 获取缺失行数
        for i in range(n):
            df2.loc[i+2] = ["~\\" for i in range(df2_cols)]
        return df1, df2
    elif df1_rows < df2_rows:
        n = df2_rows - df1_rows
        for i in range(n):
            df1.loc[i+2] = ["~\\" for i in range(df1_cols)]
        return df1, df2
    else:
        return df1,df2


def set_cellcolor(wb, sheet, fill_list, path1, path2, fgColor="FFFF0000"):
    """
    设置单元格背景颜色
    :param wb: openpyxl.workbook对象
    :param sheet:Worksheet 对象
    :param fill_list:填充单元格集
    :param path1:主文件路径
    :param path2:副文件路径
    :param fgColor:背景颜色
    :return:None
    """
    fill = PatternFill("solid", fgColor=fgColor)  # 设置样式
    for cell in fill_list:
        sheet[cell].fill = fill  # 遍历集合填充
    if check_PermissionError(path1) and check_PermissionError(path2):  # 判断是否打开
        wb.save(path1)  # 保存文件


def set_acellcolor(wb, sheet, cell, path1, path2, fgColor="FFFF0000"):
    """
    设置单元格背景颜色
    :param wb: openpyxl.workbook对象
    :param sheet:Worksheet 对象
    :param cell:准备填充的单个单元格
    :param path1:主文件路径
    :param path2:副文件路径
    :param fgColor:背景颜色
    :return:None
    """
    fill = PatternFill("solid", fgColor=fgColor)  # 设置样式
    sheet[cell].fill = fill
    if check_PermissionError(path1) and check_PermissionError(path2):  # 判断是否打开
        wb.save(path1)  # 保存文件


def dudge_merged(cell):
    """
    返回某个合并范围内的所有单元格
    :param cell:合并范围
    :return: cells_list：拆解的合并单元格坐标
    """

    # if type(cell).__name__ == 'MergedCell':
    #     print("TRue")

    cells_list = []  # 存储单元格坐标

    lst = cell.split(':')
    start, end = lst[0], lst[-1]  # 获取头字母，尾字母
    start_letter, start_number = ord(start[0]), int(start[1:])  # 转换为ASCII码
    end_letter, end_number = ord(end[0]), int(end[1:])  # 获取头数字，尾数字
    if start_letter == end_letter:  # 判断是否为横向合并
        for i in range(end_number - start_number + 1):  # 根据头尾数字差 获取列数
            cells_list.append(chr(start_letter) + str(start_number + i))
    else:
        col = column_index_from_string(end[0]) - column_index_from_string(start[0]) + 1  # 计算列数
        row = end_number - start_number + 1  # 计算行数
        col_list = [chr(start_letter + i) for i in range(col)]  # 生成范围内的所有列字母
        row_list = [start_number + i for i in range(row)]  # 生成范围内的所有行号
        for c in col_list:
            for r in row_list:
                cells_list.append(c + str(r))
    return cells_list


def get_merage_list(merage):
    """
    返回所有合并单元格范围内的所有单元格
    :param merage: 合并单元格合集
    :return: 扩充的单元格

    get_merage_list('A1:A3') --> return: [['A1','A2','A3']]
    get_merage_list('A1:B2') --> return: [['A1','A2','B1','B2']]
    """
    merged_cells = []
    if len(merage) > 0:
        for range_cell in merage:
            values = dudge_merged(str(range_cell))
            merged_cells.append(values)
    else:
        values = dudge_merged(str(merage[0]))
        merged_cells.append(values)
    return merged_cells



