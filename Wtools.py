#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import openpyxl
import sys
import WtoolsFuc
import pandas as pd
import traceback
import webbrowser
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QWidget, QFileDialog
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QMessageBox
from operator import eq
from time import time

error_report = ""


class Ui_MainWindow(QWidget):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(623, 417)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("other.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout4 = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout4.setObjectName("gridLayout_4")
        self.gridLayout3 = QtWidgets.QGridLayout()
        self.gridLayout3.setObjectName("gridLayout_3")
        self.gridLayout2 = QtWidgets.QGridLayout()
        self.gridLayout2.setObjectName("gridLayout_2")
        self.label = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Agency FB")
        font.setPointSize(22)
        self.label.setFont(font)
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.gridLayout2.addWidget(self.label, 0, 0, 1, 1)
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.pushButton2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton2.setObjectName("pushButton_2")
        self.gridLayout.addWidget(self.pushButton2, 1, 0, 1, 1)
        self.lineEdit2 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit2.setObjectName("lineEdit_2")
        self.gridLayout.addWidget(self.lineEdit2, 1, 1, 1, 1)
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setObjectName("pushButton")
        self.gridLayout.addWidget(self.pushButton, 0, 0, 1, 1)
        self.comboBox2 = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox2.setObjectName("comboBox_2")
        self.gridLayout.addWidget(self.comboBox2, 1, 2, 1, 1)
        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setObjectName("comboBox")
        self.gridLayout.addWidget(self.comboBox, 0, 2, 1, 1)
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setObjectName("lineEdit")
        self.gridLayout.addWidget(self.lineEdit, 0, 1, 1, 1)
        self.gridLayout2.addLayout(self.gridLayout, 1, 0, 1, 1)
        self.gridLayout3.addLayout(self.gridLayout2, 0, 0, 1, 1)
        self.progressBar = QtWidgets.QProgressBar(self.centralwidget)
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")
        self.gridLayout3.addWidget(self.progressBar, 1, 0, 1, 1)
        self.gridLayout4.addLayout(self.gridLayout3, 0, 0, 1, 3)
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setText("")
        self.label_3.setObjectName("label_3")
        self.gridLayout4.addWidget(self.label_3, 1, 0, 1, 1)
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setText("")
        self.label_2.setObjectName("label_2")
        self.gridLayout4.addWidget(self.label_2, 1, 1, 1, 1)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.pushButton3 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton3.setObjectName("pushButton_3")
        self.horizontalLayout.addWidget(self.pushButton3)
        self.pushButton4 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton4.setObjectName("pushButton_4")
        self.horizontalLayout.addWidget(self.pushButton4)
        self.gridLayout4.addLayout(self.horizontalLayout, 1, 2, 1, 1)
        self.textEdit = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit.setObjectName("textEdit")
        self.gridLayout4.addWidget(self.textEdit, 2, 0, 1, 3)
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.menuBar = QtWidgets.QMenuBar(MainWindow)
        self.menuBar.setGeometry(QtCore.QRect(0, 0, 623, 23))
        self.menuBar.setObjectName("menuBar")
        self.menu = QtWidgets.QMenu(self.menuBar)
        self.menu.setObjectName("menu")
        self.menu_2 = QtWidgets.QMenu(self.menuBar)
        self.menu_2.setObjectName("menu_2")
        MainWindow.setMenuBar(self.menuBar)
        self.action = QtWidgets.QAction(MainWindow)
        self.action.setObjectName("action")
        self.action_4 = QtWidgets.QAction(MainWindow)
        self.action_4.setObjectName("action_4")
        self.action_C = QtWidgets.QAction(MainWindow)
        self.action_C.setObjectName("action_C")
        self.action_R = QtWidgets.QAction(MainWindow)
        self.action_R.setObjectName("action_R")
        self.action_F = QtWidgets.QAction(MainWindow)
        self.action_F.setObjectName("action_F")
        self.action_B = QtWidgets.QAction(MainWindow)
        self.action_B.setObjectName("action_B")
        self.action_W = QtWidgets.QAction(MainWindow)
        self.action_W.setObjectName("action_W")
        self.menu.addAction(self.action_4)
        self.menu_2.addAction(self.action_C)
        self.menu_2.addAction(self.action_R)
        self.menu_2.addAction(self.action_F)
        self.menu_2.addAction(self.action_B)
        self.menu_2.addSeparator()
        self.menu_2.addAction(self.action_W)
        self.menuBar.addAction(self.menu_2.menuAction())
        self.menuBar.addAction(self.menu.menuAction())

        self.retranslateUi(MainWindow)
        self.pushButton.clicked.connect(self.open_master_file)
        self.pushButton2.clicked.connect(self.open_other_file)
        self.pushButton4.clicked.connect(self.check_excel)
        self.pushButton3.clicked.connect(self.clear_lineedit)
        self.lineEdit.textChanged.connect(self.master_combobox)
        self.lineEdit2.textChanged.connect(self.other_combobox)
        self.action_B.triggered.connect(self.check_excel)
        self.action_C.triggered.connect(self.clear_lineedit)
        self.action_F.triggered.connect(self.refresh)
        self.action_W.triggered.connect(MainWindow.close)
        self.action_R.triggered.connect(self.reset)
        self.action_4.triggered.connect(self.help)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Wtools"))
        self.label.setText(_translate("MainWindow", "Excel内容对比"))
        self.pushButton2.setText(_translate("MainWindow", "选择副文件"))
        self.pushButton.setText(_translate("MainWindow", "选择主文件"))
        self.pushButton3.setText(_translate("MainWindow", "清除"))
        self.pushButton4.setText(_translate("MainWindow", "开始对比"))
        self.menu.setTitle(_translate("MainWindow", "帮助"))
        self.menu_2.setTitle(_translate("MainWindow", "文件"))
        self.action.setText(_translate("MainWindow", "帮助"))
        self.action_4.setText(_translate("MainWindow", "帮助(&H)"))
        self.action_C.setText(_translate("MainWindow", "清除(&C)"))
        self.action_R.setText(_translate("MainWindow", "重置(&R)"))
        self.action_F.setText(_translate("MainWindow", "刷新(&F)"))
        self.action_B.setText(_translate("MainWindow", "开始对比(&B)"))
        self.action_W.setText(_translate("MainWindow", "关闭(&W)"))

    def open_master_file(self):
        """
        打开主文件
        :return: None
        QFileDialog.getOpenFileName(self, 文件选择框标题, 打开的默认路径, 过滤文件后缀)
        """
        self.lineEdit.setText(QFileDialog.getOpenFileName(self, "选择主文件", "", "Excel Files(*.xlsx)")[0])

    def open_other_file(self):
        """
        打开副文件
        :return: None
        """
        self.lineEdit2.setText(QFileDialog.getOpenFileName(self, "选择副文件", "", "Excel Files(*.xlsx)")[0])

    def check_excel(self):
        """
        查询数据合法性：
        1、lineEdit和lineEdie2是否为空r
        2、传入的主副文件是否被打开
        :return:None
        """
        if self.lineEdit.text() and self.lineEdit2.text():  # 判断是否为真
            Ui_MainWindow.master_path = self.lineEdit.text()  # 通过类属性来传递主文件路径
            Ui_MainWindow.other_path = self.lineEdit2.text()  # 通过类属性来传递副文件路径
            if WtoolsFuc.check_PermissionError(Ui_MainWindow.master_path) and WtoolsFuc.check_PermissionError(
                    Ui_MainWindow.other_path):
                self.excel_operation(Ui_MainWindow.master_path, Ui_MainWindow.other_path)
            else:
                self.critical("请关闭主/副文件")
                QtWidgets.QApplication.processEvents()  # 刷新窗体
        else:
            self.critical("请选择文件")

    def excel_operation(self, path1, path2):
        """
        数据处理的主要函数
        :param path1:主文件路径
        :param path2: 副文件路径
        :return: None
        """
        try:
            start_time = time()  # 程序计时
            Ui_MainWindow.master_sheet = self.comboBox.currentIndex()  # 通过类属性传递主副文件的sheet表index
            Ui_MainWindow.other_sheet = self.comboBox2.currentIndex()

            master_df = pd.read_excel(path1, header=None, sheet_name=Ui_MainWindow.master_sheet).fillna(
                "~\\")  # 读取主文件生成DataFrame
            other_df = pd.read_excel(path2, header=None, sheet_name=Ui_MainWindow.other_sheet).fillna(
                "~\\")  # 读取副文件生成DataFrame
            master_wb, master_sheet, merage = WtoolsFuc.get_max_rows(path1, path2,
                                                                       Ui_MainWindow.master_sheet)  # 获取主文件Workbook、Worksheet对象

            master_rows, master_cols = master_df.shape  # 获取主文件的总行数
            other_rows, other_cols = other_df.shape  # 获取副文件的总行数
            max_rows = max(master_rows, other_rows)  # 获取最大行数

            if master_rows != other_rows:
                master_df, other_df = WtoolsFuc.fill_rows(master_df, other_df,master_rows,other_rows)  # 填充行
            fill_cell = self.judege_cols(master_cols, other_cols, max_rows, master_df, other_df)

            # 异同单元格判断
            falsecell_count = len(fill_cell)  # 获取异同单元格数量
            merage_count = 0  # 合并单元格数

            """
            判断异同单元格的是否为合并单元格
            """
            if falsecell_count > 0:
                if len(merage) > 0:
                    merged_cells = WtoolsFuc.get_merage_list(merage)  # 返回文中合并单元格的具体范围
                    for merged_cell in merged_cells:
                        for fi_cell in fill_cell:
                            if fi_cell in merged_cell:
                                merage_count += 1
                                WtoolsFuc.set_cellcolor(master_wb, master_sheet, merged_cell, path1, path2,
                                                          fgColor="1874CD")
                            else:
                                WtoolsFuc.set_acellcolor(master_wb, master_sheet, fi_cell, path1, path2)
                else:
                    WtoolsFuc.set_cellcolor(master_wb, master_sheet, fill_cell, path1, path2)

            self.progressBar.setValue(100)  # 设置进度条进度为100
            if falsecell_count == 0:
                self.textEdit.append("excel文件操作成功，没有发现错误...")
            else:
                self.textEdit.append("excel文件操作成功，共发现{}处错误，已进行高亮处理，请打开主文件进行查看...".format(falsecell_count))
            if merage_count > 0:
                self.textEdit.append("程序发现excel有{}处合并单元格范围内存在异值，已进行蓝色背景处理，请打开主文件进行查看...".format(merage_count))
            end_time = time()
            self.textEdit.append("累计耗时：{}秒".format(end_time - start_time))
            self.information("excel内容对比完成！")
        except IndexError:
            self.textEdit.append(traceback.format_exc())
            self.critical("文件第{}行首单元格是否为合并单元格？请还原单元格".format(Ui_MainWindow.index_error))
        except ValueError:
            traceback.print_exc()
            self.textEdit.append(traceback.format_exc())
            self.critical("异同过大，是否选错主副表？")
        except BaseException:
            # traceback.print_exc()
            self.textEdit.append(traceback.format_exc())
            self.textEdit.append("\n")

    def excel_io(self, rows, master_df, other_df):
        """
        excel的补全，判断DateFrame行数是否一致
        :param rows: 文件的总行数
        :param master_df: 主文件的DateFrame
        :param other_df: 副文件的DateFrame
        :return: 需要填充的单元格集合
        """
        global index_error
        fill_cell = []
        for i in range(rows):
            master_list = list(master_df.values[i])  # 读取一行
            other_list = list(other_df.values[i])
            if not eq(master_list, other_list):
                master_list, other_list, lenl = WtoolsFuc.fill_cols(master_list, other_list)
                for n in range(lenl):
                    if not eq(master_list[n], other_list[n]):
                        fill_cell.append(get_column_letter(n + 1) + str(i + 1))
        return fill_cell

    def judege_cols(self, master_cols, other_cols, max_rows, wf1, wf2):
        """
        判断文件列数是否相同
        :param master_cols: 主文件列数
        :param other_cols: 副文件列数
        :param max_rows: 最大列数
        :param wf1: 主文件对象
        :param wf2: 副文件对象
        :return: 异同单元格合集
        """
        fell_list = []  # 异同单元格合集
        for i in range(max_rows):
            master_lst, other_lst = [], []
            try:
                master_lst = list(wf1.values[i])
            except IndexError:
                master_lst = (["~\\" for i in range(other_cols)])
            try:
                other_lst = list(wf2.values[i])
            except IndexError:
                other_lst = (["~\\" for i in range(master_cols)])
            if len(master_lst) != len(other_lst):
                master_lst, other_lst, lenl = WtoolsFuc.fill_cols(master_lst, other_lst)
                if not eq(master_lst,other_lst):
                    for n in range(lenl):
                        if not eq(master_lst[n],other_lst[n]):
                            fell_list.append(get_column_letter(n+1)+str(i+1))
        return fell_list

    def critical(self, t):
        """
        弹出警告窗口
        :param t: 自定义消息
        :return: None
        """
        QMessageBox.critical(None, "错误", t, QMessageBox.Ok)

    def information(self, t):
        """
        弹出提示窗口
        :param t: 自定义消息
        :return: None
        """
        QMessageBox.information(self, "提示", t, QMessageBox.Yes, QMessageBox.Yes)

    def clear_lineedit(self):
        """
        清除状态栏信息、进度条
        :return: None
        """
        if self.textEdit.toPlainText():
            self.textEdit.clear()
            self.progressBar.setValue(0)

    def master_combobox(self):
        """
        设置下拉列表框值
        :return: None
        """
        if self.comboBox.count() > 0:
            self.comboBox.clear()
        wb = openpyxl.load_workbook(self.lineEdit.text(), read_only=True)
        self.comboBox.addItems(wb.sheetnames)

    def other_combobox(self):
        """
        设置下拉列表框值
        :return: None
        """
        if self.comboBox2.count() > 0:
            self.comboBox2.clear()
        wb = openpyxl.load_workbook(self.lineEdit2.text(), read_only=True)
        self.comboBox2.addItems(wb.sheetnames)

    def refresh(self):
        """
        刷新窗体
        :return: None
        """
        QtWidgets.QApplication.processEvents()

    def reset(self):
        """
        重置所有控件
        :return: None
        """
        self.clear_lineedit()
        self.progressBar.setValue(0)
        self.textEdit.setText("")

    def help(self):
        """
        打开说明文档
        :return: None
        """
        html = "help.html"
        if html in os.listdir("./"):
            webbrowser.open("help.html")
        else:
            self.critical("没有找到帮助文件！")


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
